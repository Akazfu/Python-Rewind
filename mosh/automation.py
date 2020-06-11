import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_wb(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    # cell = sheet.cell(1, 1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        new_cell = sheet.cell(row, 4)
        new_cell.value = corrected_price
        cell_e = sheet.cell(row, 5)
        cell_e.value = cell.value + new_cell.value

    values = Reference(sheet, 
                    min_row = 2, 
                    max_row = sheet.max_row,
                    min_col = 4,
                    max_col = 5)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a9')

    wb.save('transactions2.xlsx')

filename = input('Which file you want to process? ')
process_wb(filename)